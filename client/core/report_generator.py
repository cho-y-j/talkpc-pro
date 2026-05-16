"""
ReportGenerator - 발송 결과 리포트 생성 모듈
성공/실패 통계, 이력 관리, 엑셀 내보내기
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


class ReportGenerator:
    """발송 리포트 생성기"""

    def __init__(self, log_dir: str = None):
        self.log_dir = Path(log_dir) if log_dir else Path("./logs")
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.current_session = []
        self.session_start = None

    def start_session(self):
        """새 발송 세션 시작"""
        self.current_session = []
        self.session_start = datetime.now()

    def add_result(self, result_dict: dict):
        """발송 결과 추가"""
        self.current_session.append({
            **result_dict,
            "recorded_at": datetime.now().isoformat()
        })

    def get_statistics(self) -> dict:
        """현재 세션 통계"""
        total = len(self.current_session)
        success = sum(1 for r in self.current_session if r["status"] == "success")
        failed = sum(1 for r in self.current_session if r["status"] != "success" and r["status"] != "skipped")
        skipped = sum(1 for r in self.current_session if r["status"] == "skipped")

        return {
            "total": total,
            "success": success,
            "failed": failed,
            "skipped": skipped,
            "success_rate": round(success / total * 100, 1) if total > 0 else 0,
            "duration": str(datetime.now() - self.session_start) if self.session_start else "0:00:00",
            "session_start": self.session_start.isoformat() if self.session_start else None
        }

    def get_failed_list(self) -> list[dict]:
        """실패 목록 반환"""
        return [
            r for r in self.current_session
            if r["status"] not in ("success", "skipped")
        ]

    def get_success_list(self) -> list[dict]:
        """성공 목록 반환"""
        return [r for r in self.current_session if r["status"] == "success"]

    def save_session_log(self) -> str:
        """세션 로그 JSON 저장"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.log_dir / f"session_{timestamp}.json"

        log_data = {
            "session_start": self.session_start.isoformat() if self.session_start else None,
            "session_end": datetime.now().isoformat(),
            "statistics": self.get_statistics(),
            "results": self.current_session
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(log_data, f, ensure_ascii=False, indent=2)

        return str(filepath)

    def export_report_excel(self, filepath: str = None) -> str:
        """발송 리포트 엑셀 내보내기"""
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = str(self.log_dir / f"report_{timestamp}.xlsx")

        wb = openpyxl.Workbook()

        # -- 요약 시트 --
        ws_summary = wb.active
        ws_summary.title = "발송 요약"

        stats = self.get_statistics()

        # 스타일
        header_font = Font(bold=True, size=14)
        stat_font = Font(size=12)
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws_summary["A1"] = "KakaoTalk 발송 리포트"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A3"] = "발송 일시"
        ws_summary["B3"] = stats.get("session_start", "")
        ws_summary["A4"] = "소요 시간"
        ws_summary["B4"] = stats["duration"]
        ws_summary["A6"] = "총 발송"
        ws_summary["B6"] = stats["total"]
        ws_summary["A7"] = "성공"
        ws_summary["B7"] = stats["success"]
        ws_summary["B7"].fill = success_fill
        ws_summary["A8"] = "실패"
        ws_summary["B8"] = stats["failed"]
        ws_summary["B8"].fill = fail_fill
        ws_summary["A9"] = "건너뜀"
        ws_summary["B9"] = stats["skipped"]
        ws_summary["A10"] = "성공률"
        ws_summary["B10"] = f"{stats['success_rate']}%"

        ws_summary.column_dimensions["A"].width = 15
        ws_summary.column_dimensions["B"].width = 25

        # -- 상세 시트 --
        ws_detail = wb.create_sheet("상세 결과")
        headers = ["이름", "상태", "메시지 (일부)", "상세", "시간"]
        ws_detail.append(headers)

        for cell in ws_detail[1]:
            cell.font = Font(bold=True)

        status_map = {
            "success": "✅ 성공",
            "not_found": "❌ 검색실패",
            "ocr_error": "❌ OCR오류",
            "send_error": "❌ 전송실패",
            "skipped": "⏭ 건너뜀"
        }

        for r in self.current_session:
            ws_detail.append([
                r.get("contact_name", ""),
                status_map.get(r.get("status", ""), r.get("status", "")),
                r.get("message", ""),
                r.get("detail", ""),
                r.get("recorded_at", "")
            ])

        # 열 너비
        for i, w in enumerate([15, 12, 30, 25, 22]):
            ws_detail.column_dimensions[chr(65 + i)].width = w

        # -- 실패 목록 시트 --
        failed = self.get_failed_list()
        if failed:
            ws_failed = wb.create_sheet("실패 목록")
            ws_failed.append(["이름", "실패 원인", "상세"])
            for cell in ws_failed[1]:
                cell.font = Font(bold=True)
            for r in failed:
                ws_failed.append([
                    r.get("contact_name", ""),
                    r.get("status", ""),
                    r.get("detail", "")
                ])

        wb.save(filepath)
        return filepath

    def get_history(self, limit: int = 10) -> list[dict]:
        """최근 발송 이력 로드"""
        log_files = sorted(self.log_dir.glob("session_*.json"), reverse=True)[:limit]
        history = []
        for fp in log_files:
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                history.append({
                    "file": fp.name,
                    "start": data.get("session_start"),
                    "statistics": data.get("statistics", {})
                })
            except Exception:
                continue
        return history
