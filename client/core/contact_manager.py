"""
ContactManager - 연락처 관리 모듈
엑셀 업로드, 카테고리 관리, CRUD
"""

import json
import os
import quopri
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _decode_vcf_value(raw_value: str, params: dict) -> str:
    """VCF 속성값 디코드 (Quoted-Printable / charset 처리)"""
    encoding = params.get("ENCODING", "").upper()
    charset = params.get("CHARSET", "UTF-8") or "UTF-8"
    if encoding == "QUOTED-PRINTABLE":
        try:
            decoded_bytes = quopri.decodestring(raw_value)
            return decoded_bytes.decode(charset, errors="replace")
        except (LookupError, ValueError):
            return raw_value
    return raw_value


def _normalize_phone(raw: str) -> str:
    """전화번호 정리 (공백/괄호 제거, +82 → 0)"""
    s = raw.strip()
    s = re.sub(r"[\s()]", "", s)
    if s.startswith("+82"):
        s = "0" + s[3:].lstrip("-")
    return s


class Contact:
    """연락처 데이터 클래스"""

    def __init__(self, name: str, category: str = "other", phone: str = "",
                 company: str = "", position: str = "", memo: str = "",
                 birthday: str = "", anniversary: str = "",
                 contact_id: str = None):
        self.id = contact_id or f"{name}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        self.name = name
        self.category = category
        self.phone = phone
        self.company = company
        self.position = position
        self.memo = memo
        self.birthday = birthday      # MM-DD 형식
        self.anniversary = anniversary  # MM-DD 형식
        self.created_at = datetime.now().isoformat()
        self.last_sent = None
        self.send_count = 0

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "name": self.name,
            "category": self.category,
            "phone": self.phone,
            "company": self.company,
            "position": self.position,
            "memo": self.memo,
            "birthday": self.birthday,
            "anniversary": self.anniversary,
            "created_at": self.created_at,
            "last_sent": self.last_sent,
            "send_count": self.send_count
        }

    @classmethod
    def from_dict(cls, data: dict) -> "Contact":
        contact = cls(
            name=data["name"],
            category=data.get("category", "other"),
            phone=data.get("phone", ""),
            company=data.get("company", ""),
            position=data.get("position", ""),
            memo=data.get("memo", ""),
            birthday=data.get("birthday", ""),
            anniversary=data.get("anniversary", ""),
            contact_id=data.get("id")
        )
        contact.created_at = data.get("created_at", contact.created_at)
        contact.last_sent = data.get("last_sent")
        contact.send_count = data.get("send_count", 0)
        return contact


class ContactManager:
    """연락처 CRUD 및 관리"""

    DEFAULT_CATEGORIES = ["friend", "family", "business", "vip", "kakao_friend", "other"]
    # 절대 삭제 불가 — kakao_friend 는 자동 동기화 기능과 결합
    PROTECTED_CATEGORIES = {"kakao_friend"}

    def __init__(self, data_path: str = None):
        self.data_path = Path(data_path) if data_path else Path("./data/contacts.json")
        self.contacts: list[Contact] = []
        self.custom_categories: list[str] = []
        # 사용자가 숨긴(=삭제한) 기본 카테고리. UI 에서 제외하지만 contact 데이터는 보존.
        self.hidden_default_categories: list[str] = []
        self.load()

    def load(self):
        """JSON 파일에서 연락처 로드"""
        if self.data_path.exists():
            try:
                with open(self.data_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.contacts = [
                    Contact.from_dict(c) for c in data.get("contacts", [])
                ]
                self.custom_categories = data.get("custom_categories", [])
                self.hidden_default_categories = data.get("hidden_default_categories", [])
            except (json.JSONDecodeError, KeyError):
                self.contacts = []

    def save(self):
        """연락처를 JSON 파일로 저장"""
        self.data_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "contacts": [c.to_dict() for c in self.contacts],
            "custom_categories": self.custom_categories,
            "hidden_default_categories": self.hidden_default_categories,
            "last_updated": datetime.now().isoformat()
        }
        with open(self.data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def add(self, contact: Contact) -> bool:
        """연락처 추가"""
        # 중복 확인 (이름 기준)
        for c in self.contacts:
            if c.name == contact.name and c.category == contact.category:
                return False
        self.contacts.append(contact)
        self.save()
        return True

    def update(self, contact_id: str, save=True, **kwargs) -> bool:
        """연락처 수정"""
        for c in self.contacts:
            if c.id == contact_id:
                for key, value in kwargs.items():
                    if hasattr(c, key):
                        setattr(c, key, value)
                if save:
                    self.save()
                return True
        return False

    def batch_update_category(self, contact_ids: list, category: str):
        """여러 연락처 카테고리 일괄 변경 (한 번만 저장)"""
        id_set = set(contact_ids)
        for c in self.contacts:
            if c.id in id_set:
                c.category = category
        self.save()

    def delete(self, contact_id: str) -> bool:
        """연락처 삭제"""
        before = len(self.contacts)
        self.contacts = [c for c in self.contacts if c.id != contact_id]
        if len(self.contacts) < before:
            self.save()
            return True
        return False

    def get_by_category(self, category: str) -> list[Contact]:
        """카테고리별 연락처 조회"""
        if category == "all":
            return self.contacts
        return [c for c in self.contacts if c.category == category]

    def get_by_name(self, name: str) -> Optional[Contact]:
        """이름으로 연락처 검색"""
        for c in self.contacts:
            if c.name == name:
                return c
        return None

    def search(self, query: str) -> list[Contact]:
        """연락처 검색 (이름, 회사, 메모, 생일).

        생일 검색 지원:
          - "03-15" / "03" 처럼 부분 문자열 매칭
          - "3월" / "03월" → 해당 월(MM) 생일자
        """
        q = query.lower().strip()
        # "N월" → 생일 월(MM)
        month_mm = None
        m = re.match(r"(\d{1,2})\s*월$", q)
        if m:
            month_mm = f"{int(m.group(1)):02d}"
        result = []
        for c in self.contacts:
            bday = c.birthday or ""
            if (q in c.name.lower() or q in c.company.lower()
                    or q in c.memo.lower() or (q and q in bday)):
                result.append(c)
            elif month_mm and bday[:2] == month_mm:
                result.append(c)
        return result

    def get_birthdays_today(self) -> list[Contact]:
        """오늘 생일인 연락처"""
        today = datetime.now().strftime("%m-%d")
        return [c for c in self.contacts if (c.birthday or "") == today]

    def get_birthdays_this_month(self) -> list[Contact]:
        """이번 달 생일인 연락처 (일자 오름차순)"""
        mm = datetime.now().strftime("%m")
        result = [c for c in self.contacts if (c.birthday or "")[:2] == mm]
        result.sort(key=lambda c: c.birthday)
        return result

    def get_all(self) -> list[Contact]:
        """전체 연락처 조회"""
        return self.contacts

    def get_count(self) -> int:
        """전체 연락처 수"""
        return len(self.contacts)

    def get_category_counts(self) -> dict:
        """카테고리별 연락처 수"""
        counts = {}
        for c in self.contacts:
            counts[c.category] = counts.get(c.category, 0) + 1
        return counts

    def import_from_excel(self, filepath: str, default_category: str = None) -> dict:
        """
        엑셀 파일에서 연락처 일괄 가져오기

        엑셀 형식:
        | 이름 | 카테고리 | 전화번호 | 회사 | 직급 | 메모 |

        Returns:
            {"success": int, "skipped": int, "errors": []}
        """
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        result = {"success": 0, "skipped": 0, "errors": []}

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active

            headers = []
            header_map = {
                "이름": "name", "name": "name",
                "카테고리": "category", "category": "category",
                "전화번호": "phone", "phone": "phone", "연락처": "phone",
                "회사": "company", "company": "company",
                "직급": "position", "position": "position",
                "메모": "memo", "memo": "memo", "비고": "memo",
                "생일": "birthday", "birthday": "birthday",
                "기념일": "anniversary", "anniversary": "anniversary",
            }

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 헤더 행 파싱
                    headers = [
                        header_map.get(str(cell).strip().lower() if cell else "", "")
                        for cell in row
                    ]
                    continue

                if not row[0]:  # 이름이 없으면 건너뛰기
                    continue

                try:
                    data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers) and headers[col_idx]:
                            data[headers[col_idx]] = str(cell).strip() if cell else ""

                    if "name" not in data or not data["name"]:
                        result["skipped"] += 1
                        continue

                    # 카테고리: 엑셀에 있으면 사용, 없으면 default_category, 그것도 없으면 "미지정"
                    cat = data.get("category", "").strip()
                    if not cat:
                        cat = default_category or "미지정"
                    contact = Contact(
                        name=data.get("name", ""),
                        category=cat,
                        phone=data.get("phone", ""),
                        company=data.get("company", ""),
                        position=data.get("position", ""),
                        memo=data.get("memo", ""),
                        birthday=data.get("birthday", ""),
                        anniversary=data.get("anniversary", ""),
                    )

                    if self.add(contact):
                        result["success"] += 1
                    else:
                        result["skipped"] += 1

                except Exception as e:
                    result["errors"].append(f"행 {row_idx + 1}: {str(e)}")

            wb.close()

        except Exception as e:
            result["errors"].append(f"파일 읽기 오류: {str(e)}")

        return result

    def import_from_vcf(self, filepath: str, default_category: str = None) -> dict:
        """
        VCF (vCard) 파일에서 연락처 일괄 가져오기

        지원:
        - 다중 카드 (BEGIN:VCARD ... END:VCARD 반복)
        - UTF-8 / CP949 인코딩 자동 폴백
        - Quoted-Printable 인코딩 (삼성폰 한글)
        - 라인 폴딩 (이어지는 줄이 공백/탭으로 시작)
        - TEL 우선순위: CELL/MOBILE > HOME > WORK > 기타

        Returns:
            {"success": int, "skipped": int, "errors": []}
        """
        result = {"success": 0, "skipped": 0, "errors": []}

        # 파일 읽기 + 인코딩 폴백
        try:
            with open(filepath, "rb") as f:
                raw_bytes = f.read()
        except Exception as e:
            result["errors"].append(f"파일 읽기 오류: {e}")
            return result

        text = None
        for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
            try:
                text = raw_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = raw_bytes.decode("utf-8", errors="replace")

        text = text.replace("\r\n", "\n").replace("\r", "\n")
        raw_lines = text.split("\n")

        # 1단계: 라인 언폴딩 (공백/탭으로 시작하는 줄은 이전 줄 연속)
        unfolded = []
        for line in raw_lines:
            if line.startswith((" ", "\t")) and unfolded:
                unfolded[-1] += line[1:]
            else:
                unfolded.append(line)

        # 2단계: VCARD 블록 분리
        cards = []
        current = []
        in_card = False
        for line in unfolded:
            stripped = line.strip()
            upper = stripped.upper()
            if upper == "BEGIN:VCARD":
                in_card = True
                current = []
            elif upper == "END:VCARD":
                if in_card and current:
                    cards.append(current)
                in_card = False
                current = []
            elif in_card:
                current.append(line)

        # 3단계: 각 카드 파싱
        for idx, card_lines in enumerate(cards):
            try:
                # QP 소프트 라인 브레이크 병합 (값이 = 로 끝나면 다음 줄 이어짐)
                merged = []
                i = 0
                while i < len(card_lines):
                    line = card_lines[i]
                    if "QUOTED-PRINTABLE" in line.upper() and line.rstrip().endswith("="):
                        combined = line.rstrip()[:-1]
                        i += 1
                        while i < len(card_lines):
                            nxt = card_lines[i]
                            if nxt.rstrip().endswith("="):
                                combined += nxt.rstrip()[:-1]
                                i += 1
                            else:
                                combined += nxt
                                i += 1
                                break
                        merged.append(combined)
                    else:
                        merged.append(line)
                        i += 1

                name = ""
                fn_name = ""
                phones = []  # (priority, number)

                for prop_line in merged:
                    if ":" not in prop_line:
                        continue
                    key_part, _, value = prop_line.partition(":")
                    segments = key_part.split(";")
                    prop = segments[0].upper().strip()
                    params = {}
                    type_values = []
                    for seg in segments[1:]:
                        if "=" in seg:
                            k, _, v = seg.partition("=")
                            k_up = k.strip().upper()
                            if k_up == "TYPE":
                                for tv in v.split(","):
                                    type_values.append(tv.strip().upper())
                            else:
                                params[k_up] = v.strip()
                        else:
                            type_values.append(seg.strip().upper())

                    decoded = _decode_vcf_value(value, params)

                    if prop == "FN":
                        if not fn_name:
                            fn_name = decoded.strip()
                    elif prop == "N":
                        # N: family;given;additional;prefix;suffix
                        parts = decoded.split(";")
                        family = parts[0].strip() if len(parts) > 0 else ""
                        given = parts[1].strip() if len(parts) > 1 else ""
                        # 한글은 성+이름 붙여쓰기, 영어는 given family 순서
                        if family and given:
                            if re.search(r"[가-힣]", family + given):
                                name = family + given
                            else:
                                name = f"{given} {family}"
                        else:
                            name = (family or given).strip()
                    elif prop == "TEL":
                        types_combined = " ".join(type_values).upper()
                        if "CELL" in types_combined or "MOBILE" in types_combined:
                            priority = 0
                        elif "HOME" in types_combined:
                            priority = 1
                        elif "WORK" in types_combined:
                            priority = 2
                        else:
                            priority = 3
                        phone = _normalize_phone(decoded)
                        if phone:
                            phones.append((priority, phone))

                final_name = name or fn_name
                if not final_name:
                    result["skipped"] += 1
                    continue

                phone_str = ""
                if phones:
                    phones.sort(key=lambda x: x[0])
                    phone_str = phones[0][1]

                cat = default_category or "미지정"
                contact = Contact(
                    name=final_name,
                    category=cat,
                    phone=phone_str,
                )

                if self.add(contact):
                    result["success"] += 1
                else:
                    result["skipped"] += 1

            except Exception as e:
                result["errors"].append(f"카드 {idx + 1}: {e}")

        return result

    def export_to_excel(self, filepath: str, contacts: list[Contact] = None) -> bool:
        """연락처를 엑셀 파일로 내보내기"""
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        contacts = contacts or self.contacts

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "연락처"

        # 헤더
        headers = ["이름", "카테고리", "전화번호", "회사", "직급", "메모",
                   "생일", "기념일", "마지막 발송", "발송 횟수"]
        ws.append(headers)

        # 데이터
        for c in contacts:
            ws.append([
                c.name, c.category, c.phone, c.company,
                c.position, c.memo, c.birthday or "", c.anniversary or "",
                c.last_sent or "", c.send_count
            ])

        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

        wb.save(filepath)
        return True

    def get_all_categories(self) -> list[str]:
        """기본(숨긴 것 제외) + 커스텀 + 실제 사용 중인 모든 카테고리 반환"""
        used = set(c.category for c in self.contacts)
        hidden = set(self.hidden_default_categories)
        all_cats = [c for c in self.DEFAULT_CATEGORIES if c not in hidden]
        for cat in self.custom_categories:
            if cat not in all_cats:
                all_cats.append(cat)
        # 사용 중인 카테고리는 숨겨도 표시 (사용자 데이터 보존)
        for cat in used:
            if cat not in all_cats:
                all_cats.append(cat)
        return all_cats

    def add_category(self, category: str) -> bool:
        """커스텀 카테고리 추가"""
        category = category.strip()
        if not category:
            return False
        # 숨겨진 기본 카테고리 복원 처리
        if category in self.hidden_default_categories:
            self.hidden_default_categories.remove(category)
            self.save()
            return True
        if category in self.DEFAULT_CATEGORIES or category in self.custom_categories:
            return False
        self.custom_categories.append(category)
        self.save()
        return True

    def delete_category(self, category: str) -> bool:
        """카테고리 삭제 — kakao_friend 만 보호. 기본/커스텀 모두 삭제 가능.

        - 커스텀: custom_categories 에서 제거
        - 기본: hidden_default_categories 에 추가 (소프트 삭제, 데이터 보존)
        """
        if category in self.PROTECTED_CATEGORIES:
            return False
        if category in self.custom_categories:
            self.custom_categories.remove(category)
            self.save()
            return True
        if category in self.DEFAULT_CATEGORIES:
            if category not in self.hidden_default_categories:
                self.hidden_default_categories.append(category)
                self.save()
                return True
            return False
        return False

    def create_sample_excel(self, filepath: str) -> bool:
        """샘플 엑셀 파일 생성 (가져오기 형식 안내)"""
        if openpyxl is None:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "연락처"

        # 헤더
        headers = ["이름", "카테고리", "전화번호", "회사", "직급", "메모", "생일", "기념일"]
        ws.append(headers)

        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill, Alignment
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 샘플 데이터
        samples = [
            ["홍길동", "friend", "010-1234-5678", "ABC회사", "대리", "동창", "03-15", ""],
            ["김영희", "business", "010-9876-5432", "XYZ건설", "팀장", "거래처 담당자", "07-22", "05-10"],
            ["이철수", "vip", "010-5555-1234", "DEF기획", "대표", "주요 고객", "", "12-25"],
            ["박지영", "family", "010-3333-4444", "", "", "사촌", "11-03", ""],
        ]
        for sample in samples:
            ws.append(sample)

        # 열 너비
        widths = [12, 12, 16, 16, 10, 20, 10, 10]
        for i, w in enumerate(widths, 1):
            col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
            ws.column_dimensions[col_letter].width = w

        # 안내 시트
        ws2 = wb.create_sheet("안내")
        ws2.append(["카카오톡 자동 발송 - 연락처 가져오기 안내"])
        ws2.append([])
        ws2.append(["필수 항목: 이름 (첫 번째 열)"])
        ws2.append(["선택 항목: 카테고리, 전화번호, 회사, 직급, 메모, 생일, 기념일"])
        ws2.append([])
        ws2.append(["생일/기념일 형식: MM-DD (예: 03-15, 12-25)"])
        ws2.append([])
        ws2.append(["카테고리 종류:"])
        all_cats = self.get_all_categories()
        ws2.append([", ".join(all_cats)])
        ws2.append([])
        ws2.append(["* 첫 번째 행은 반드시 헤더여야 합니다"])
        ws2.append(["* 이름이 비어있는 행은 건너뜁니다"])
        ws2.append(["* 동일 이름+카테고리 중복은 자동 건너뜁니다"])

        wb.save(filepath)
        return True

    def mark_sent(self, contact_id: str):
        """발송 완료 마킹"""
        for c in self.contacts:
            if c.id == contact_id:
                c.last_sent = datetime.now().isoformat()
                c.send_count += 1
                self.save()
                break

    # ── 스마트 필터링 ──

    def get_unsent(self, category: str = "all") -> list[Contact]:
        """한 번도 발송하지 않은 연락처"""
        contacts = self.get_by_category(category)
        return [c for c in contacts if not c.last_sent]

    def get_not_sent_today(self, category: str = "all") -> list[Contact]:
        """오늘 아직 안 보낸 연락처"""
        today = datetime.now().strftime("%Y-%m-%d")
        contacts = self.get_by_category(category)
        return [
            c for c in contacts
            if not c.last_sent or not c.last_sent.startswith(today)
        ]

    def get_not_sent_within_days(self, days: int, category: str = "all") -> list[Contact]:
        """N일 이내에 안 보낸 연락처 (재발송 방지)"""
        from datetime import timedelta
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        contacts = self.get_by_category(category)
        return [
            c for c in contacts
            if not c.last_sent or c.last_sent < cutoff
        ]

    def get_sent_today(self, category: str = "all") -> list[Contact]:
        """오늘 이미 보낸 연락처"""
        today = datetime.now().strftime("%Y-%m-%d")
        contacts = self.get_by_category(category)
        return [
            c for c in contacts
            if c.last_sent and c.last_sent.startswith(today)
        ]

    def get_send_stats(self) -> dict:
        """발송 통계 요약"""
        today = datetime.now().strftime("%Y-%m-%d")
        total = len(self.contacts)
        sent_today = sum(1 for c in self.contacts
                         if c.last_sent and c.last_sent.startswith(today))
        never_sent = sum(1 for c in self.contacts if not c.last_sent)
        total_sends = sum(c.send_count for c in self.contacts)

        return {
            "total_contacts": total,
            "sent_today": sent_today,
            "remaining_today": total - sent_today,
            "never_sent": never_sent,
            "total_sends": total_sends,
        }
